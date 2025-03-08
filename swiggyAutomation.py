import json
import os
import pandas as pd
import datetime
import requests
import re
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from openpyxl.utils import get_column_letter



# Load search queries from Food_Keywords.xlsx
keywords_df = pd.read_excel("Book2.xlsx", header=None)
search_queries = keywords_df.iloc[:, 0].dropna().tolist()  # Extracts column values (ignoring header)
# Load locations from Locations.xlsx
locations_df = pd.read_excel("Book1.xlsx")
locations = locations_df.to_dict(orient="records")  # Convert rows to list of dictionaries

# Set up output folder and file naming
save_path = "D:/SwiggyData/"  # Modify this to your preferred folder
os.makedirs(save_path, exist_ok=True)

timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
# output_excel_path = os.path.join(save_path, f"SwiggyData-{timestamp}.xlsx")
output_excel_path = f"SwiggyData-{timestamp}.xlsx"  # Save locally first

# Create an Excel writer for storing results
excel_writer = pd.ExcelWriter(output_excel_path, engine="openpyxl")

# Function to set fixed column widths
def set_fixed_column_width(writer, sheet_name):
    """Sets a fixed width for specific columns in the Excel sheet"""
    sheet = writer.sheets[sheet_name]  # Get the worksheet

    fixed_widths = {  # Column width mapping (adjust as needed)
        "A": 25,  # Dish Name
        "B": 10,  # Rating
        "C": 25,  # Restaurant Name
        "D": 10,  # Total Ratings
        "E": 10,  # Price
        "F": 20,  # Locality
        "G": 18,  # Category
        "H": 60,  # Description (Wider for readability)
        "I": 20,  # Area Name
        "J": 30,  # Cuisine
        "K": 15,  # Discount
        "L": 20,  # Discount Details
        "M": 15,  # Discount Type
    }

    for col_letter, width in fixed_widths.items():
        sheet.column_dimensions[col_letter].width = width

# Auto-adjust column width function
def adjust_column_width(writer, sheet_name):
    """Auto-adjusts column widths based on text length"""
    sheet = writer.sheets[sheet_name]  # Get the worksheet

    for col_idx, col_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Set column width (increase for readability)
        if col_letter == "D":  # "Description" column (adjust as per your column index)
            sheet.column_dimensions[col_letter].width = max_length * 3  # 3x wider
        else:
            sheet.column_dimensions[col_letter].width = max_length + 5  # Adjust general columns


def upload_to_drive(file_path, folder_id):
    """Uploads the file to Google Drive in a specific folder."""
    SCOPES = ["https://www.googleapis.com/auth/drive.file"]
    SERVICE_ACCOUNT_FILE = "credentials.json"  # Path to your service account JSON file

    # Authenticate with Google Drive API
    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES
    )
    service = build("drive", "v3", credentials=credentials)

    # File metadata
    file_metadata = {
        "name": os.path.basename(file_path),
        "parents": [folder_id],  # Google Drive Folder ID
    }
    
    media = MediaFileUpload(file_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Upload file
    file = service.files().create(body=file_metadata, media_body=media, fields="id").execute()
    print(f"✅ File uploaded successfully: {file_path} (File ID: {file.get('id')})")

# Function to clean text fields to remove illegal characters
def clean_text(text):
    """Removes illegal characters that Excel cannot handle."""
    if isinstance(text, str):
        return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text)  # Removes control characters
    return text

# Function to merge all location data and remove duplicates before saving
def merge_and_remove_duplicates(data_list):
    """Merges all data from multiple locations and removes duplicate dish-restaurant combinations."""
    if not data_list:
        return pd.DataFrame()  # Return empty DataFrame if no data is present

    merged_df = pd.concat(data_list, ignore_index=True)  # Merge all location data
    cleaned_df = merged_df.drop_duplicates(subset=["Dish Name", "Restaurant Name"], keep="first")  # Remove duplicates
    return cleaned_df

def convert_total_ratings(rating_str):
    """
    Converts "K+" formatted ratings into numerical values.
    Example: "1.3K+" → 1300, "10K+" → 10000, "123" → 123
    """
    if isinstance(rating_str, str):
        rating_str = rating_str.replace("+", "").strip()  # Remove "+"
        
        # Convert "K" values (e.g., "1.3K" -> 1300)
        match = re.match(r"([\d.]+)K", rating_str, re.IGNORECASE)
        if match:
            return int(float(match.group(1)) * 1000)

        # Convert "M" values if any (e.g., "2.5M" -> 2,500,000)
        match = re.match(r"([\d.]+)M", rating_str, re.IGNORECASE)
        if match:
            return int(float(match.group(1)) * 1_000_000)

        # Convert normal numbers
        try:
            return int(rating_str)
        except ValueError:
            return 0  # Default to 0 if unknown format
    return 0  # Default for non-string values

# Function to fetch Swiggy API data
def fetch_swiggy_data(lat, lng, query):
    """Fetches and processes data from Swiggy API"""
    
    # Construct API URL correctly (fixing spaces)
    api_url = f"https://www.swiggy.com/dapi/restaurants/search/v3?lat={str(lat).strip()}&lng={str(lng).strip()}&str={query.strip().replace(' ', '%20')}&trackingId=undefined&submitAction=ENTER"

    # Set up request headers to avoid blocking
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }

    # Make request
    response = requests.get(api_url, headers=headers)
    
    if response.status_code == 200:
        return response.json()
    else:
        print(f"❌ Failed to fetch data for {query} (Status code: {response.status_code})")
        return None

# Loop through each keyword first
for query in search_queries:
    print(f"\n🔍 Searching for: {query}")

    location_data = []  # Stores cleaned results across all locations

    # Loop through all locations for this keyword
    for location in locations:
        lat = location["Latitude"]
        lng = location["Longitude"]

        print(f"   📍 Checking location: Latitude {lat}, Longitude {lng}")

        # Fetch API Data
        data = fetch_swiggy_data(lat, lng, query)
        if not data:
            continue  # Skip if no data

        dish_data = []
        try:
            cards = data.get("data", {}).get("cards", [])

            # Extract "DISH" data correctly
            dish_cards = []
            for card in cards:
                if "groupedCard" in card:
                    dish_cards = card["groupedCard"].get("cardGroupMap", {}).get("DISH", {}).get("cards", [])
                    break  # Stop loop once we find 'groupedCard'

            # Extract dish details
            for dish_card in dish_cards:
                dish_info = dish_card.get("card", {}).get("card", {}).get("info", {})
                restaurant_info = dish_card.get("card", {}).get("card", {}).get("restaurant", {}).get("info", {})

                if dish_info:
                    dish_data.append({
                        "Dish Name": dish_info.get("name", "N/A"),
                        "Rating": dish_info.get("ratings", {}).get("aggregatedRating", {}).get("rating", "N/A"),
                        "Restaurant Name": restaurant_info.get("name", "N/A"),
                        "Total Ratings": convert_total_ratings(restaurant_info.get("totalRatingsString", "0")),
                        "Price (₹)": dish_info.get("price", 0) / 100,  # Convert paise to ₹
                        "Locality": restaurant_info.get("locality", "N/A"),
                        "Category": dish_info.get("category", "N/A"),
                        "costForTwoMessage": restaurant_info.get("costForTwoMessage", "N/A"),
                        "Description": dish_info.get("description", "N/A"),
                        "Area Name": restaurant_info.get("areaName", "N/A"),
                        "Cuisine": ", ".join(restaurant_info.get("cuisines", [])),
                       "Discount": restaurant_info.get("aggregatedDiscountInfoV3", {}).get("header", "N/A"),
                        "Discount Details": restaurant_info.get("aggregatedDiscountInfoV3", {}).get("subHeader", "N/A"),
                        "Discount Type": restaurant_info.get("aggregatedDiscountInfoV3", {}).get("discountTag", "N/A"),
                    })

        except Exception as e:
            print(f"⚠️ Error extracting data for {query}: {e}")

        # Append data from this location
        if dish_data:
            location_data.append(pd.DataFrame(dish_data))

    # Merge all location data before deduplication
    df_final = merge_and_remove_duplicates(location_data)

    # Save cleaned data
    if not df_final.empty:
        df_final = df_final.map(clean_text)  # Clean text for Excel compatibility
        sheet_name = query[:31]  # Excel sheet name limit is 31 characters
        df_final.to_excel(excel_writer, sheet_name=sheet_name, index=False)

        # Auto-adjust column width
        set_fixed_column_width(excel_writer, sheet_name)
    else:
        print(f"   ❌ No data found for {query}")

# Save Excel
GOOGLE_DRIVE_FOLDER_ID = "1gmh07ZHRImVHe-icxgeJryV7w3SKPNYK"  # Replace with actual folder ID from Google Drive

# Save Excel File
if len(excel_writer.book.sheetnames) > 0:
    excel_writer.close()
    print(f"✅ Excel file saved locally: {output_excel_path}")
    upload_to_drive(output_excel_path, GOOGLE_DRIVE_FOLDER_ID)  # Upload to Google Drive
else:
    print("⚠️ No data found. Skipping upload.")